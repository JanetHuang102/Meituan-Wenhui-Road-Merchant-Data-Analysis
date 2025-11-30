import pandas as pd
import os
import jieba
from snownlp import SnowNLP
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# 定义文本清理函数
def clean_text(text):
    if pd.isna(text):
        return ""
    text = str(text)
    # 去除方括号标记的表情符号
    text = re.sub(r'\[.*?\]', '', text)
    # 保留中文、常见标点、空格、字母数字
    text = re.sub(r'[^\u4e00-\u9fa5，。！？；：\s\w]', '', text)
    return text.strip()


# 定义分词函数
def segment_text(text):
    if text == "":
        return ""
    words = jieba.cut(text)
    return ' '.join(words)


# 定义情感分析函数
def sentiment_analysis(text):
    if text == "":
        return None
    s = SnowNLP(text)
    return s.sentiments


# 定义计算字符串宽度的函数（用于Excel列宽调整）
def get_string_width(text):
    width = 0
    for char in str(text):
        if '\u4e00' <= char <= '\u9fff':  # 中文字符
            width += 2
        else:
            width += 1
    return width


# 定义自动调整列宽的函数
def auto_adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_width = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_width = get_string_width(cell.value)
                if cell_width > max_width:
                    max_width = cell_width
        # 设置列宽：最大宽度加边距，限制最大宽度为50
        adjusted_width = min(max_width + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width


# 输入和输出目录
input_dir = r"D:\Desktop\爬虫\美团\清洗后"
output_dir = r"D:\Desktop\爬虫\美团\分析后"
os.makedirs(output_dir, exist_ok=True)

# 遍历输入目录中的所有Excel文件
for filename in os.listdir(input_dir):
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        input_path = os.path.join(input_dir, filename)
        base, ext = os.path.splitext(filename)
        output_path = os.path.join(output_dir, base + '.xlsx')  # 统一输出为.xlsx格式

        # 根据文件扩展名选择引擎读取文件
        if ext == '.xlsx':
            df = pd.read_excel(input_path, engine='openpyxl')
        else:  # .xls文件
            df = pd.read_excel(input_path, engine='xlrd')

        # 检查是否存在评论内容列
        if '评论内容' not in df.columns:
            print(f"警告：文件 {filename} 中没有找到'评论内容'列，跳过处理。")
            continue

        # 文本清理、分词和情感分析
        df['清理后评论'] = df['评论内容'].apply(clean_text)
        df['分词结果'] = df['清理后评论'].apply(segment_text)
        df['情感得分'] = df['清理后评论'].apply(sentiment_analysis)

        # 将情感得分转换为字符串，保留4位小数
        df['情感得分'] = df['情感得分'].apply(lambda x: f"{x:.20f}" if x is not None else "")

        # 保存为.xlsx文件
        df.to_excel(output_path, index=False, engine='openpyxl')

        # 使用openpyxl调整列宽
        wb = load_workbook(output_path)
        ws = wb.active
        auto_adjust_column_width(ws)
        wb.save(output_path)

        print(f"处理完成：{filename}，输出到 {output_path}")

print("所有文件处理完毕！")